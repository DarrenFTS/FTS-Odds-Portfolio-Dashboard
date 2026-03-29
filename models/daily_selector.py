"""
Daily Bet Selector - Clean Version

This module generates daily betting shortlists
"""

import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models.value_calculator import ValueCalculator
from systems.all_systems import get_system


class DailyBetSelector:
    """Generate daily betting selections"""
    
    def __init__(self, config_dir='config'):
        """Initialize the daily bet selector"""
        self.config_dir = Path(config_dir)
        self.value_calc = ValueCalculator(config_dir)
        
        # Initialize all systems
        self.systems = {
            'Home Win': get_system('Home Win', config_dir),
            'O2.5 Back': get_system('O2.5 Back', config_dir),
            'O3.5 Lay': get_system('O3.5 Lay', config_dir),
            'U1.5 Lay': get_system('U1.5 Lay', config_dir),
            'FHGU0.5 Lay': get_system('FHGU0.5 Lay', config_dir)
        }
    
    def load_fixtures(self, fixtures_file, target_date=None):
        """Load daily fixtures from file"""
        
        print(f"Loading fixtures from: {fixtures_file}")
        
        # Load Excel file with header in row 2
        try:
            fixtures = pd.read_excel(fixtures_file, header=1, engine='openpyxl')
            print(f"  Loaded {len(fixtures)} rows")
        except Exception as e:
            raise ValueError(f"Could not read file: {str(e)}")
        
        # Clean column names
        fixtures.columns = [str(col).strip() for col in fixtures.columns]
        
        # Convert Date column
        if 'Date' not in fixtures.columns:
            raise ValueError(f"No 'Date' column found")
        
        fixtures['Date'] = pd.to_datetime(fixtures['Date'], errors='coerce')
        fixtures = fixtures[fixtures['Date'].notna()]
        
        # Filter by date if specified
        if target_date:
            if isinstance(target_date, str):
                target_date = pd.to_datetime(target_date)
            
            # Handle date comparison
            if hasattr(target_date, 'date'):
                target_date_obj = target_date.date()
            else:
                target_date_obj = target_date
            
            fixtures = fixtures[fixtures['Date'].dt.date == target_date_obj]
            print(f"  Filtered to {len(fixtures)} fixtures for {target_date_obj}")
        
        # Standardize league column
        if 'Competition' in fixtures.columns:
            fixtures['League'] = fixtures['Competition']
        elif 'League' not in fixtures.columns:
            raise ValueError("No 'League' or 'Competition' column")
        
        print(f"  Ready: {len(fixtures)} fixtures loaded")
        
        return fixtures
    
    def generate_selections(self, fixtures_file, target_date=None):
        """Generate betting selections"""
        
        fixtures = self.load_fixtures(fixtures_file, target_date)
        
        if len(fixtures) == 0:
            return pd.DataFrame()
        
        print(f"\nScanning {len(fixtures)} fixtures...")
        
        all_bets = []
        
        for system_name, system in self.systems.items():
            signals = system.scan_fixtures(fixtures)
            
            for signal in signals:
                home_odds = signal.get('home_odds') if system_name == 'O2.5 Back' else None
                
                value_result = self.value_calc.calculate_value_score(
                    system=system_name,
                    league=signal['league'],
                    odds=signal['odds'],
                    config=signal['config'],
                    home_odds=home_odds
                )
                
                if value_result['score'] == 0:
                    continue
                
                bet = {
                    'Date': signal['date'],
                    'Time': signal['time'],
                    'League': signal['league'],
                    'Home Team': signal['home_team'],
                    'Away Team': signal['away_team'],
                    'System': system_name,
                    'Odds': signal['odds'],
                    'Odds Range': signal['odds_range'],
                    'Value Score': value_result['score'],
                    'Confidence': value_result['confidence'],
                    'In Exact Range': value_result['in_exact_range'],
                    'Filter Passed': value_result['filter_passed']
                }
                
                if system_name == 'O2.5 Back':
                    bet['Home Odds'] = signal['home_odds']
                    if signal['filter_passed']:
                        bet['Filter Status'] = "Home > 2.00 ✅"
                    elif signal['filter_in_buffer']:
                        bet['Filter Status'] = "Home 1.80-1.99 ⚠️ (buffer)"
                    else:
                        bet['Filter Status'] = ""
                else:
                    bet['Home Odds'] = None
                    bet['Filter Status'] = ""
                
                bet['ROI %'] = value_result['breakdown']['roi_pct']
                bet['Sample Size'] = value_result['breakdown']['sample_size']
                
                all_bets.append(bet)
        
        if len(all_bets) == 0:
            print("\nNo qualifying bets found")
            return pd.DataFrame()
        
        df = pd.DataFrame(all_bets)
        df = df.sort_values(['Value Score', 'Time'], ascending=[False, True])
        
        print(f"\nFound {len(df)} qualifying bets:")
        print(f"  HIGH: {len(df[df['Confidence'] == 'HIGH'])}")
        print(f"  SPECULATIVE: {len(df[df['Confidence'] == 'SPECULATIVE'])}")
        
        return df
    
    def export_to_excel(self, selections_df, output_file):
        """Export to Excel"""
        
        if len(selections_df) == 0:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Today's Shortlist"
        
        colors = {'HIGH': 'C6EFCE', 'SPECULATIVE': 'FFC7CE'}
        
        output_cols = [
            'Date', 'Time', 'League', 'Home Team', 'Away Team',
            'System', 'Odds Range', 'Odds', 'Confidence', 'Value Score',
            'Filter Status', 'Notes'
        ]
        
        all_systems = ['Home Win', 'O2.5 Back', 'U1.5 Lay', 'O3.5 Lay', 'FHGU0.5 Lay']
        output_cols.extend(all_systems)
        output_cols.extend(['P/L', 'Cum P/L'])
        
        output_rows = []
        for _, bet in selections_df.iterrows():
            row = {
                'Date': bet['Date'].strftime('%Y-%m-%d') if isinstance(bet['Date'], pd.Timestamp) else bet['Date'],
                'Time': bet['Time'],
                'League': bet['League'],
                'Home Team': bet['Home Team'],
                'Away Team': bet['Away Team'],
                'System': bet['System'],
                'Odds Range': bet['Odds Range'],
                'Odds': bet['Odds'],
                'Confidence': bet['Confidence'],
                'Value Score': bet['Value Score'],
                'Filter Status': bet.get('Filter Status', ''),
                'Notes': 'Monitor closer to kick-off' if bet['Confidence'] == 'SPECULATIVE' else ''
            }
            
            for sys in all_systems:
                row[sys] = np.nan
            
            row['P/L'] = np.nan
            row['Cum P/L'] = np.nan
            
            output_rows.append(row)
        
        for col_idx, header in enumerate(output_cols, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        for row_idx, row_data in enumerate(output_rows, 2):
            confidence = row_data['Confidence']
            row_color = colors.get(confidence, 'FFFFFF')
            
            for col_idx, col_name in enumerate(output_cols, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col_name)
                
                if col_name in ['Odds', 'Value Score'] and not pd.isna(value):
                    cell.value = float(value)
                    cell.number_format = '0.0'
                else:
                    cell.value = value
                
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
                
                if col_name in ['Odds', 'Value Score'] or col_name in all_systems:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_name in ['Confidence']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        widths = {'A': 12, 'B': 8, 'C': 25, 'D': 20, 'E': 20, 'F': 18, 
                  'G': 12, 'H': 8, 'I': 12, 'J': 12, 'K': 25, 'L': 40}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        for i in range(len(all_systems) + 2):
            ws.column_dimensions[chr(77 + i)].width = 13
        
        wb.save(output_file)
        print(f"\nExported to: {output_file}")
    
    def print_summary(self, selections_df):
        """Print summary"""
        if len(selections_df) == 0:
            return
        
        print("\n" + "="*80)
        print("DAILY SELECTIONS SUMMARY")
        print("="*80)
        print(f"\nTotal Bets: {len(selections_df)}")
        print(f"  HIGH: {len(selections_df[selections_df['Confidence'] == 'HIGH'])}")
        print(f"  SPECULATIVE: {len(selections_df[selections_df['Confidence'] == 'SPECULATIVE'])}")
        print("\nBy System:")
        for system in selections_df['System'].value_counts().index:
            count = len(selections_df[selections_df['System'] == system])
            high = len(selections_df[(selections_df['System'] == system) & 
                                    (selections_df['Confidence'] == 'HIGH')])
            print(f"  {system}: {count} total ({high} HIGH)")
