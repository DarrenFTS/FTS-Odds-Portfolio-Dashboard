"""
Enhanced Daily Bet Selector - Fully Integrated with Value Scoring

QUALIFYING ODDS FEATURE:
Value scores are calculated using "qualifying odds" rather than current odds.
For bets in the buffer zone:
- If current odds are BELOW exact range: Use exact_min (entry point when rising)
- If current odds are ABOVE exact range: Use exact_max (entry point when falling)
- If in exact range: Use current odds

This ensures value scores represent the actual value at the point where bets qualify,
not misleading scores from buffer zone odds.

CRITICAL FIXES INCLUDED:
1. Strict buffer range validation - bets outside buffer_max are REJECTED
2. Daily fixtures column mapping - handles both file formats
3. Proper LAY/BACK bet type detection
4. Value score calculation with qualifying odds
5. Only shows bets with Value Score 50+ (configurable)

This version includes ALL value scoring logic built-in.
No external dependencies on value_scoring.py needed.
"""

import pandas as pd
import numpy as np
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from systems.all_systems import get_system
except ImportError:
    from all_systems import get_system


class EnhancedDailySelector:
    """Generate daily betting selections with integrated advanced value scoring"""
    
    def __init__(self, config_dir='config'):
        self.config_dir = Path(config_dir)
        
        # Load portfolio stats for scoring
        with open(self.config_dir / 'portfolio_stats.json', 'r') as f:
            data = json.load(f)
            self.portfolio_stats = data['stats']
            self.max_roi = data['max_roi']
        
        # Build strike rate lookup
        self.strike_rates = {}
        for key, stats in self.portfolio_stats.items():
            self.strike_rates[key] = {
                'strike_rate': stats['strike_rate'] / 100,
                'sample_size': stats['total_bets']
            }
        
        # Initialize all systems
        self.systems = {
            'Home Win': get_system('Home Win', config_dir),
            'O2.5 Back': get_system('O2.5 Back', config_dir),
            'O3.5 Lay': get_system('O3.5 Lay', config_dir),
            'U1.5 Lay': get_system('U1.5 Lay', config_dir),
            'FHGU0.5 Lay': get_system('FHGU0.5 Lay', config_dir)
        }
    
    def calculate_model_probability(self, system, league):
        """Get model probability from historical strike rate"""
        key = f"{system}|{league}"
        if key not in self.strike_rates:
            return None
        
        strike_rate = self.strike_rates[key]['strike_rate']
        
        # For all systems: model probability = strike rate
        # Strike rate already represents probability that OUR bet wins
        # - Home Win SR 65% = 65% chance home wins
        # - O2.5 Back SR 58% = 58% chance over 2.5 happens
        # - O3.5 Lay SR 85% = 85% chance UNDER 3.5 happens (over doesn't)
        # - U1.5 Lay SR 88% = 88% chance OVER 1.5 happens (under doesn't)
        # - FHGU0.5 Lay SR 86% = 86% chance NO first half goal (goal doesn't happen)
        
        return strike_rate
    
    def calculate_expected_value(self, model_prob, odds, bet_type='back'):
        """Calculate EV%"""
        if model_prob is None:
            return 0
        
        if bet_type == 'back':
            ev = (model_prob * (odds - 1)) + ((1 - model_prob) * (-1))
        else:  # lay
            ev = (model_prob * 1) + ((1 - model_prob) * (-(odds - 1)))
        
        return ev * 100
    
    def calculate_sample_reliability(self, sample_size):
        """
        Calculate reliability score based on sample size
        
        New penalty structure:
        - 100+ bets: No penalty (100% of points)
        - 50-99 bets: -10% penalty (90% of points)
        - 30-49 bets: -20% penalty (80% of points)
        - <30 bets: -30% penalty (70% of points)
        """
        if sample_size <= 0:
            return 0
        
        # Base reliability using log scale (maxes at 200 bets)
        base_reliability = min(np.log10(sample_size) / np.log10(200), 1.0)
        
        # Apply sample size penalty
        if sample_size >= 100:
            penalty_multiplier = 1.0  # No penalty
        elif sample_size >= 50:
            penalty_multiplier = 0.9  # -10% penalty
        elif sample_size >= 30:
            penalty_multiplier = 0.8  # -20% penalty
        else:  # < 30 bets
            penalty_multiplier = 0.7  # -30% penalty
        
        return base_reliability * penalty_multiplier
    
    def get_qualifying_odds(self, current_odds, exact_min, exact_max):
        """
        Returns the odds to use for value score calculation.
        
        For buffer zone bets, use the entry point odds (exact range boundary)
        rather than current odds. This gives accurate value scores at the
        point where the bet would actually qualify.
        
        Args:
            current_odds: Current market odds
            exact_min: Minimum exact range odds
            exact_max: Maximum exact range odds
            
        Returns:
            tuple: (qualifying_odds, status_message, in_exact_range)
        """
        if current_odds < exact_min:
            # In low buffer - use exact_min (entry point when rising)
            return exact_min, f"⬆ Needs to rise to {exact_min:.2f}", False
        elif current_odds > exact_max:
            # In high buffer - use exact_max (entry point when falling)
            return exact_max, f"⬇ Needs to drop to {exact_max:.2f}", False
        else:
            # In exact range - use current odds
            return current_odds, "✅ In exact range", True
    
    def calculate_value_score(self, system, league, current_odds, exact_min=0, exact_max=999, in_exact_range=True, filter_passed=True):
        """
        Calculate 0-100 value score using QUALIFYING odds.
        
        IMPORTANT: Uses qualifying odds (entry point) for buffer zone bets,
        not current odds. This ensures value scores represent the actual
        value when the bet qualifies.
        """
        
        key = f"{system}|{league}"
        
        if key not in self.portfolio_stats:
            return {
                'value_score': 0,
                'model_probability': None,
                'market_probability': None,
                'expected_value': 0,
                'historical_roi': 0,
                'sample_size': 0,
                'interpretation': 'No data',
                'qualifying_odds': current_odds,
                'odds_status': 'No data'
            }
        
        stats = self.portfolio_stats[key]
        bet_type = 'lay' if 'Lay' in system else 'back'
        
        # Get qualifying odds (entry point if in buffer, current if in exact range)
        qualifying_odds, odds_status, is_in_exact = self.get_qualifying_odds(
            current_odds, exact_min, exact_max
        )
        
        # Calculate probabilities using QUALIFYING odds
        model_prob = self.calculate_model_probability(system, league)
        
        # Market probability based on QUALIFYING odds (not current odds)
        # This ensures we calculate value at the entry point
        if bet_type == 'back':
            market_prob = 1 / qualifying_odds
        else:  # lay
            market_prob = 1 - (1 / qualifying_odds)
        
        # Expected value using QUALIFYING odds
        ev_pct = self.calculate_expected_value(model_prob, qualifying_odds, bet_type)
        
        # Component 1: Expected Value (30 points)
        if ev_pct >= 50:
            ev_points = 30
        elif ev_pct >= 0:
            ev_points = 15 + (ev_pct / 50) * 15
        else:
            ev_points = max(0, 15 + (ev_pct / 50) * 15)
        
        # Component 2: Historical ROI (25 points)
        system_roi = stats['roi']
        roi_normalized = min(system_roi / self.max_roi, 1.0)
        roi_points = roi_normalized * 25
        
        # Component 3: Sample Reliability (20 points)
        sample_size = stats['total_bets']
        reliability = self.calculate_sample_reliability(sample_size)
        sample_points = reliability * 20
        
        # Component 4: League Performance (15 points)
        league_points = roi_normalized * 15
        
        # Component 5: Odds Band (10 points)
        odds_band_points = roi_normalized * 10
        
        # Calculate total
        total_points = ev_points + roi_points + sample_points + league_points + odds_band_points
        
        # NO PENALTIES APPLIED
        # Bets only appear if they qualify, so no need to reduce their value
        # - Removed: Buffer zone penalty (was -15%)
        # - Removed: O2.5 Back filter penalty (was -10%)
        
        final_score = min(total_points, 100)
        
        # Interpretation
        if final_score >= 80:
            interpretation = "Exceptional - Strong edge"
        elif final_score >= 70:
            interpretation = "Excellent - Very strong"
        elif final_score >= 60:
            interpretation = "Very Good - Clear value"
        elif final_score >= 50:
            interpretation = "Good - Solid opportunity"
        elif final_score >= 40:
            interpretation = "Fair - Moderate value"
        else:
            interpretation = "Marginal - Low confidence"
        
        return {
            'value_score': round(final_score, 1),
            'model_probability': round(model_prob * 100, 2) if model_prob else None,
            'market_probability': round(market_prob * 100, 2),
            'expected_value': round(ev_pct, 2),
            'historical_roi': round(system_roi, 2),
            'sample_size': sample_size,
            'interpretation': interpretation,
            'qualifying_odds': qualifying_odds,
            'current_odds': current_odds,
            'odds_status': odds_status,
            'in_exact_range': is_in_exact
        }
    
    def load_fixtures(self, fixtures_file, target_date=None):
        """Load daily fixtures with automatic column mapping"""
        
        try:
            fixtures = pd.read_excel(fixtures_file, header=1, engine='openpyxl')
        except Exception as e:
            raise ValueError(f"Could not read file: {str(e)}")
        
        # Clean columns
        fixtures.columns = [str(col).strip() for col in fixtures.columns]
        
        # Date handling
        if 'Date' not in fixtures.columns:
            raise ValueError("No 'Date' column found")
        
        fixtures['Date'] = pd.to_datetime(fixtures['Date'], errors='coerce')
        fixtures = fixtures[fixtures['Date'].notna()]
        
        # Filter by date
        if target_date:
            if isinstance(target_date, str):
                target_date = pd.to_datetime(target_date)
            
            if hasattr(target_date, 'date'):
                target_date_obj = target_date.date()
            else:
                target_date_obj = target_date
            
            fixtures = fixtures[fixtures['Date'].dt.date == target_date_obj]
        
        # League column
        if 'Competition' in fixtures.columns:
            fixtures['League'] = fixtures['Competition']
        elif 'League' not in fixtures.columns:
            raise ValueError("No 'League' or 'Competition' column")
        
        # ✅ CRITICAL FIX: Map odds columns for daily fixtures format
        # This handles files exported from FTS which use different column names
        daily_fixtures_mapping = {
            'Home Win Back': 'Home Back Odds',
            'Over 2.5 Back': 'O2.5 Back Odds',
            'O3.5.1': 'O3.5 Lay Odds',
            'U1.5.1': 'U1.5 Lay Odds',
            'FHGU0.5.1': 'FHGU0.5 Lay Odds'
        }
        
        for old_name, new_name in daily_fixtures_mapping.items():
            if old_name in fixtures.columns and new_name not in fixtures.columns:
                fixtures[new_name] = fixtures[old_name]
        
        return fixtures
    
    def generate_selections(self, fixtures_file, target_date=None):
        """Generate selections with value scoring and strict validation"""
        
        fixtures = self.load_fixtures(fixtures_file, target_date)
        
        if len(fixtures) == 0:
            return pd.DataFrame()
        
        all_bets = []
        
        # Scan with each system
        # NOTE: The systems/base_system.py already validates buffer ranges
        # Any bet returned from system.scan_fixtures() has already passed:
        # - Buffer range validation (buffer_min <= odds <= buffer_max)
        # - Filter validation (for O2.5 Back)
        # - League configuration check
        for system_name, system in self.systems.items():
            signals = system.scan_fixtures(fixtures)
            
            for signal in signals:
                # Build bet record
                bet = {
                    'Date': signal['date'],
                    'Time': signal['time'],
                    'League': signal['league'],
                    'Home Team': signal['home_team'],
                    'Away Team': signal['away_team'],
                    'System': system_name,
                    'Odds': signal['odds'],
                    'Odds Range': signal['odds_range'],
                    'In Exact Range': signal['in_exact_range'],
                    'Filter Passed': signal.get('filter_passed', True)
                }
                
                # Filter status
                if system_name == 'O2.5 Back':
                    if signal['filter_passed']:
                        bet['Filter Status'] = "Home > 2.00 ✅"
                    elif signal.get('filter_in_buffer'):
                        bet['Filter Status'] = "Home 1.80-1.99 ⚠️"
                    else:
                        bet['Filter Status'] = ""
                else:
                    bet['Filter Status'] = ""
                
                # Get exact range from portfolio stats
                key = f"{system_name}|{signal['league']}"
                if key in self.portfolio_stats:
                    # Note: portfolio uses 'Odds_Min' and 'Odds_Max' (capital O)
                    exact_min = self.portfolio_stats[key].get('Odds_Min', 0)
                    exact_max = self.portfolio_stats[key].get('Odds_Max', 999)
                else:
                    exact_min = 0
                    exact_max = 999
                
                # Calculate value score with qualifying odds logic
                # NOTE: Calculations use qualifying odds, but only current odds shown in export
                score_result = self.calculate_value_score(
                    system=system_name,
                    league=signal['league'],
                    current_odds=signal['odds'],
                    exact_min=exact_min,
                    exact_max=exact_max,
                    in_exact_range=signal['in_exact_range'],
                    filter_passed=signal.get('filter_passed', True)
                )
                
                # Add score components
                # 'Odds' shows current market odds
                # But Value Score, EV%, Market Prob all based on qualifying odds
                bet['Odds'] = score_result['current_odds']  # Show current odds
                bet['Value Score'] = score_result['value_score']
                bet['Model Probability'] = score_result['model_probability']
                bet['Market Probability'] = score_result['market_probability']
                bet['Expected Value %'] = score_result['expected_value']
                bet['Historical ROI %'] = score_result['historical_roi']
                bet['Sample Size'] = score_result['sample_size']
                bet['Interpretation'] = score_result['interpretation']
                
                # Confidence category
                score = score_result['value_score']
                if score >= 80:
                    bet['Confidence'] = 'EXCEPTIONAL'
                elif score >= 70:
                    bet['Confidence'] = 'EXCELLENT'
                elif score >= 60:
                    bet['Confidence'] = 'HIGH'
                elif score >= 50:
                    bet['Confidence'] = 'GOOD'
                elif score >= 40:
                    bet['Confidence'] = 'FAIR'
                else:
                    bet['Confidence'] = 'MARGINAL'
                
                all_bets.append(bet)
        
        if len(all_bets) == 0:
            return pd.DataFrame()
        
        # Convert to DataFrame
        bets_df = pd.DataFrame(all_bets)
        
        # ✅ VALUE SCORE FILTER: Only show bets with Value Score 50+ (GOOD and above)
        # This removes FAIR (40-49), MARGINAL (30-39), and WEAK (0-29)
        # Includes: EXCEPTIONAL (80+), EXCELLENT (70-79), HIGH (60-69), and GOOD (50-59)
        # 
        # NOTE: To see ALL qualifying bets (including FAIR 40-49), change to:
        # bets_df = bets_df[bets_df['Value Score'] >= 40]
        # 
        # To remove filter completely and see ALL bets, comment out this line:
        bets_df = bets_df[bets_df['Value Score'] >= 50]
        
        if len(bets_df) == 0:
            return pd.DataFrame()
        
        # SORT: By Date and Time (earliest first)
        bets_df['Date'] = pd.to_datetime(bets_df['Date'])
        bets_df = bets_df.sort_values(['Date', 'Time'], ascending=[True, True])
        
        return bets_df
    
    def export_to_excel(self, selections_df, output_file):
        """Export to Excel with enhanced formatting"""
        
        if len(selections_df) == 0:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Value Ranked Bets"
        
        # Color scheme
        def get_color(score):
            if score >= 70:
                return 'C6EFCE'  # Dark green
            elif score >= 60:
                return 'D4EDDA'  # Light green
            elif score >= 50:
                return 'FFF3CD'  # Yellow
            elif score >= 40:
                return 'FFE5CC'  # Orange
            else:
                return 'FFC7CE'  # Red
        
        # Columns in requested order
        cols = [
            'Date', 'Time', 'League', 'Home Team', 'Away Team',
            'System', 'Odds', 'Odds Range', 'Filter Status',
            'Value Score', 'Expected Value %', 'Confidence',
            'Model Probability', 'Market Probability',
            'Historical ROI %', 'Sample Size',
            'Interpretation'
        ]
        
        # Headers
        for col_idx, header in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data
        for row_idx, (_, bet) in enumerate(selections_df.iterrows(), 2):
            color = get_color(bet['Value Score'])
            
            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = bet.get(col_name)
                
                # Format Date in UK format DD/MM/YYYY
                if col_name == 'Date':
                    if not pd.isna(value):
                        if hasattr(value, 'strftime'):
                            cell.value = value.strftime('%d/%m/%Y')
                        else:
                            cell.value = value
                # Format numeric columns
                elif col_name in ['Value Score', 'Expected Value %', 'Historical ROI %', 'Model Probability', 'Market Probability']:
                    if not pd.isna(value):
                        cell.value = float(value)
                        cell.number_format = '0.00'
                elif col_name == 'Odds':
                    if not pd.isna(value):
                        cell.value = float(value)
                        cell.number_format = '0.00'
                else:
                    cell.value = value
                
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center' if col_name in ['Value Score', 'Confidence'] else 'left', vertical='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 14
        ws.column_dimensions['O'].width = 14
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 40
        
        wb.save(output_file)
